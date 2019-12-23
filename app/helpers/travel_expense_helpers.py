from app.helpers.excel_helpers import setBorderToCell
from app.helpers.database_helpers import getHolidayByDate
import openpyxl as xl
import re
import pandas as pd
import datetime
from config import INTERNAL_FINANCE

# 横軸ラベル（["A","B", ... "BW","BX"]）
ALPHABET_COLUMNS = [chr(i) for i in range(65,65+26)] + ["A"+chr(i) for i in range(65,65+26)] + ["B"+chr(i) for i in range(65,65+24)] 

# 縦軸ラベル（["1","2", ... "54","55"])
NUMBER_ROWS = [str(i) for i in range(1,56)]


LEFT_BORDERS = {
    "thin":{
        "F":[[7,12],[14,45]],
        "W":[[14,45],],
        "AB":[[7,12],],
        "AD":[[14,49],],
        "AI":[[14,49],],
        "AM":[[7,12],],
        "AP":[[14,45],],
        "AW":[[14,45],],
        "BD":[[14,45],],
        "BI":[[7,12],],
        "BK":[[14,45],],
        "BN":[[14,45],],
        "BU":[[14,45],],
        "AU":[[54,54]],
        "BJ":[[14,45],[54,54]],
        "BM":[[14,45],[54,54]]
    },
    "bold":{
        "A":[[7,12],[14,49]],
        "J":[[1,3],],
        "P":[[1,3],],
        "AH":[[54,54],],
        "BX":[[7,12],[14,49],[54,54]],
    },
    "dashed":{
    },
}

TOP_BORDERS = {
    "thin":{
        9:[["F","BW"],],   
    },
    "bold":{
        1:[["J","O"],],
        4:[["J","O"],],
        7:[["A","BW"],],
        13:[["A","BW"],],
        14:[["A","BW"],],
        18:[["A","BW"],],
        22:[["A","BW"],],
        26:[["A","BW"],],
        30:[["A","BW"],],
        34:[["A","BW"],],
        38:[["A","BW"],],
        42:[["A","BW"],],
        46:[["A","BW"],],
        50:[["A","BW"],],
        54:[["AH","BW"],],
        55:[["AH","BW"],],
    },
    "dashed":{
        16:[["F","AC"],["AI","BJ"]],
        20:[["F","AC"],["AI","BJ"]],
        24:[["F","AC"],["AI","BJ"]],
        28:[["F","AC"],["AI","BJ"]],
        32:[["F","AC"],["AI","BJ"]],
        36:[["F","AC"],["AI","BJ"]],
        40:[["F","AC"],["AI","BJ"]],
        44:[["F","AC"],["AI","BJ"]],
    },
}

def setBorders(ws):

    # 大小比較のためアルファベット文字列を数値に変換する関数("A"=65, "AA"=91)
    def aton(alpha):

        n = len(alpha)
        return 65 + sum([ (ord(a)-64)*(26**(n-i-1)) for i,a in enumerate(alpha) ])

    for r in NUMBER_ROWS:
        for c in ALPHABET_COLUMNS:
            
            style = {"top":"default","left":"default","diagonal":False}

            # 左・上線について，線種ごとにチェックする
            for borderKind in ["thin","bold","dashed"]:

                # 左の罫線を確認
                alphabetKeys = LEFT_BORDERS[borderKind].keys()
                

                # 1. 列番号がLEFT_BORDERSのキーになっている
                # 2. まだ線種が確定していない
                # 3. 行番号がLEFT_BORDERSの値で指定した範囲に含まれる
                if c in alphabetKeys and style["left"]=="default" and sum([ start <= int(r) <= end for start,end in LEFT_BORDERS[borderKind][c] ]) > 0:
                    style["left"] = borderKind

                # 上の罫線から確認
                numberKeys = TOP_BORDERS[borderKind].keys()

                # 1. 行番号がTOP_BORDERSのキーになっている
                # 2. まだ線種が確定していない
                # 3. 列号がLEFT_BORDERSの値で指定した範囲に含まれる
                
                if int(r) in numberKeys and style["top"]=="default" and sum([ aton(start) <= aton(c) <= aton(end) for start,end in TOP_BORDERS[borderKind][int(r)] ]) > 0:
                    style["top"] = borderKind

            if c=="A" and r == "46":
                style[2]=True
            
            setBorderToCell(ws[c+r],style)



def makeExcelData(formData, userData, financeData):

    # 登録された経路数からページ数を計算
    
    n_page = (len(formData["data"])-1)//7 + 1
    
    # 固有のファイル名を取得(メールアドレスのドメインから上)
    filename = userData["e-mail"].split("@")[0]

    # シートの作製
    wb = xl.load_workbook("./app/resources/travel_expense.xlsm", read_only=False, keep_vba=True)
    ws = wb.active
    ws.title = "page0"

    # 枚数分のシートを作成
    for i in range(1,n_page):
        ws = wb.copy_worksheet(wb["page0"])
        ws.title = "page"+str(i)

    # 各ページごとに内容を
    for pageNumber in range(n_page):
        writeEachPageContents(wb, pageNumber, formData, userData, financeData)

    # データを保存
    wb.save("./app/downloads/{0}.xlsm".format(filename))
    wb.close()

    return filename

def writeEachPageContents(wb, pageNumber, formData, userData, financeData):

    ws = wb["page"+str(pageNumber)]

    # デフォルトの罫線をセット
    setBorders(ws)
    
    # 名前の設定
    ws["AM9"].value = "{0} {1}".format(userData["last-name"],userData["first-name"])
    
    # 何年何月分かの設定
    y = int(formData["year"])-2018
    ws["Y4"].value = "元" if y==1 else str(y)
    ws["AE4"].value = formData["month"]

    # 役職の設定
    ws["AB9"].value = userData["role"]

    # 所属の設定
    ws["F9"].value = userData["belonging"]

    # 取引先コードの設定
    ws["BI9"].value = userData["work-code"]


    # 共同研究者の設定
    if not formData["finance"] in INTERNAL_FINANCE:
        ws["AI46"].value = "■"
        ws["AI47"].value = "共同研究者"
    else:
        ws["AI46"].value = ""

    # 各行の設定
    for rowi, val in enumerate(formData["data"][7*pageNumber:7*(pageNumber+1)]):
        
        # 日付
        y,m,d = val["date"].split("/")   
        ws["A"  + str(18+4*rowi)] = m + " 月 " + d + " 日"
        ws["F"  + str(18+4*rowi)] = val["locate"] # 用務先
        ws["F"  + str(20+4*rowi)] = val["purpose"] # 目的
    
        # 登録経路を分割
        # 前後の移動手段がバスの場合はバス停と判断。便宜上前後に%-%を加えておく
        pattern = r"%=%|%-%"
        route_list = re.split(pattern, val["route"])
        trans_list = ["%-%",] + re.findall(pattern, val["route"]) + ["%-%",]
        
        # (バス)補正
        route_list = [ station+"(バス)" if trans_list[i]==trans_list[i+1]=="%-%" else station for i, station in enumerate(route_list) ]        
        ws["W"  + str(18+4*rowi)] = route_list[0] # 出発地
        ws["W"  + str(20+4*rowi)] = route_list[-1] # 到着地
        
        # 備考・運賃
        fare = int(val["fare"][:-5])
        round_trip = val["fare"][-3:-1]       
        ws["BN" + str(18+4*rowi)] = round_trip + "(" + "-".join(route_list) + ")"        
        ws["AD" + str(18+4*rowi)] = fare

        # 財源        
        for alpha in ("AI","AP","AW","BD"):
            for top_bottom in range(2):
                ws[alpha + str(18 + 2*top_bottom + 4*rowi)].value = financeData[top_bottom][alpha]
    
    # 最終日から日付等を設定
    y,m,d = formData["data"][-1]["date"].split("/")
    ws["BL2"].value = ("令和 " + "元" if y=="2019" else str(int(y)-2018)) + "年 " + m + "月 " + d + "日"

    # 担当
    # if userData["role"] == "学部4年":
    #     ws["BM54"] = "庵（64643)"
    # else:
    ws["BM54"] = "本木（64646)"


def finalCheck(formData):

    comment = "データを作成しました\n"
    comment += "--------------------------------------\n"

    # 月の初めに作成していた場合は年月が合っているか確認する
    if formData["day"] < 10:
        comment += "・年月分が合っているか確認してください\n"

    # 各径路にたいするチェック
    screen_shot_check = []
    for data in formData["data"]:

        y,m,d = data["date"].split("/")
        
        # 土日の確認
        if datetime.datetime.strptime(data["date"],'%Y/%m/%d').weekday() >= 5:
            comment += "・{0}/{1}は土日なので、「{0}/{1} 指導教員許可済み」とポストイットに貼って提出してください\n".format(m,d)

        # 祝日の確認
        else:

            # データベースに祝日で登録されているか確認
            match_day = getHolidayByDate(data["date"])

            if match_day != None:
                comment += "・{0}/{1}は祝日({2})なので、「{0}/{1} 指導教員許可済み」とポストイットに貼って提出してください\n".format(m,d,match_day.content)
        
        # まだスクショチェックを通過していない経路
        if not data["route"] in screen_shot_check:
            screen_shot_check.append( data["route"] )

    # 経路の確認
    for route in screen_shot_check:
        comment += "[{0}] の経路のスクリーンショットを添付してください\n".format(re.sub('%=%|%-%','-',route)) 

    return comment

