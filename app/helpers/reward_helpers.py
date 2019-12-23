import openpyxl as xl
from openpyxl.styles import Font
from app.helpers.excel_helpers import setBorderToCell

circle_pos = {

    "reward-kind":{
        0: {"left":200, "top":345, "width":30, "height":19},
        1: {"left":300, "top":345, "width":30, "height":19},
        2: {"left":380, "top":345, "width":30, "height":19},
    },

    "gender":{
        0: {"left":378, "top":425, "width":20, "height":19},
        1: {"left":408, "top":425, "width":20, "height":19},
    },

    "birthday":{
        0: {"left":385, "top":453, "width":30, "height":19},
        1: {"left":425, "top":453, "width":30, "height":19},
    },

    "term":{
        0: {"left":250, "top":623, "width":30, "height":19},
        1: {"left":370, "top":623, "width":30, "height":19},    
    },

    "unit":{
        0: {"left":400, "top":742, "width":22, "height":19},
        1: {"left":434, "top":742, "width":22, "height":19},
        2: {"left":462, "top":742, "width":22, "height":19},    
        3: {"left":490, "top":742, "width":22, "height":19},    
        4: {"left":520, "top":742, "width":22, "height":19},          
    },

    "code":{
        0: {"left":588, "top":1015, "width":30, "height":19},
        1: {"left":630, "top":1015, "width":30, "height":19},        
    },

    "travel":{
        0: {"left":275, "top":1115, "width":30, "height":19},
        # 1: {"left":350, "top":1115, "width":30, "height":19},        
    },

}

def setCircle(ws, left, top, width=30, height=19):

    img = xl.drawing.image.Image('app/resources/circle.png')
    img.drawing.left = left
    img.drawing.top = top
    img.drawing.width = width
    img.drawing.height = height    
    ws.add_image(img)

def setBorders(ws):

    style = {"top":"thin","left":"default","diagonal":False}
    
    for cell in ["G35","H35","I35","J35","K35","U37","V37","W37",]:
        setBorderToCell(ws[cell], style)


def makeExcelData(formData, userData, experimentorData):

    # シートの作製
    wb = xl.load_workbook("./app/resources/reward.xlsx", read_only=False)
    ws = wb["新領域"]

    # 罫線
    setBorders(ws)

    # 実験者の氏名
    ws["Q10"] = "{0}　{1}".format(experimentorData["last-name"],experimentorData["first-name"])
    
    # 実験者の所属
    ws["Q11"] = "{0}".format(experimentorData["belonging"])
    ws["Q11"].font = Font(size=8)

    # 報酬の種類
    setCircle(ws, **circle_pos["reward-kind"][0])

    # 実験の内容
    ws["G19"] = experimentorData["work-content"]

    # 申請者の名前
    ws["G21"] = "{0}　{1}".format(userData["last-name-reading"], userData["first-name-reading"])
    ws["G22"] = "{0}　{1}".format(userData["last-name"], userData["first-name"])
    
    # 申請者の性別
    gender = {"男":0, "女":1}[userData["gender"]]
    setCircle(ws, **circle_pos["gender"][ gender ])

    # 申請者の国籍
    ws["R21"] = userData["nationality"]

    # 申請者の誕生日
    y,m,d = userData["birthday"].split("/")
    
    birthday = 0 if int(y) < 1989 else 1
    setCircle(ws, **circle_pos["birthday"][birthday])

    ws["Q22"] = int(y) - 1996 + 8
    ws["S22"] = m
    ws["U22"] = d
    ws["P23"] = y
    ws["S23"] = m
    ws["U23"] = d

    # 申請者の取引先コード
    ws["J23"] = userData["work-code"]

    # 申請者の郵便番号
    first = formData["postal-code"][:3]
    second = formData["postal-code"][3:]
    ws["H25"] = "{0}-{1}".format(first, second)

    # 申請者の住所
    ws["K25"] = formData["address"]

    # 申請者の所属・役割
    ws["G27"] = "{0}　{1}".format(userData["belonging"], userData["role"])

    # 実施予定期間
    ys,ms,ds = experimentorData["term-start"].split("/")
    ye,me,de = experimentorData["term-end"].split("/")
    
    ws["H29"] = ys
    ws["K29"] = ms
    ws["M29"] = ds
    ws["Q29"] = ye
    ws["T29"] = me
    ws["V29"] = de    

    # 実施予定場所
    ws["G32"] = experimentorData["location"]

    # 実施予定時間
    ws["M36"] = experimentorData["hours"]
    setCircle(ws, **circle_pos["unit"][0])

    # 旅費の有無
    setCircle(ws, **circle_pos["travel"][0])

    # データを保存
    filename = userData["e-mail"].split("@")[0]
    wb.save("./app/downloads/{0}.xlsx".format(filename))
    wb.close()

    return filename